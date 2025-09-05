# -*- coding: utf-8 -*-

import numpy as np
import pandas as pd
from pathlib import Path
from tqdm.auto import tqdm
from datasets import load_dataset
from transformers import CLIPProcessor, CLIPModel
import torch
import h5py   # efficient on-disk storage

# ---- Paths ----
OUTPUT_DIR = Path("output_openai_clip")
OUTPUT_DIR.mkdir(exist_ok=True)
embeddings_path = OUTPUT_DIR / "fairface_clip_embeddings.h5"
metadata_path = OUTPUT_DIR / "fairface_metadata.csv"

# ---- Device and batch size ----
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
BATCH_SIZE = 128   # A100 can usually handle 128–256 with ViT-B/16
EMB_DIM = 512      # CLIP ViT-B/16 embedding size

# ---- CLIP Wrapper ----
class CLIP_Wrapper:
    def __init__(self, model_name="openai/clip-vit-base-patch16", device="cuda"):
        self.device = device
        self.model = CLIPModel.from_pretrained(model_name).to(device)
        self.processor = CLIPProcessor.from_pretrained(model_name)
        self.model.eval()

    def encode_images(self, pil_images):
        inputs = self.processor(images=pil_images, return_tensors="pt").to(self.device)
        with torch.no_grad():
            feats = self.model.get_image_features(**inputs)
        return feats.cpu().numpy()

# ---- Normalize embeddings ----
def normalize(x, eps=1e-12):
    norms = np.linalg.norm(x, axis=1, keepdims=True)
    return x / (norms + eps)

# ---- Stream and process dataset ----
def stream_and_save_embeddings(batch_size=BATCH_SIZE):
    # Load both subsets
    ds_125 = load_dataset("HuggingFaceM4/FairFace", "1.25")
    ds_025 = load_dataset("HuggingFaceM4/FairFace", "0.25")

    # Merge train + validation iterators
    datasets_all = [("1.25", split, ds_125[split]) for split in ds_125] + \
                   [("0.25", split, ds_025[split]) for split in ds_025]

    # Count total images
    total_images = sum(len(d) for _, _, d in datasets_all)
    print(f"✅ Total images combined: {total_images}")

    # Prepare HDF5 file
    with h5py.File(embeddings_path, "w") as f:
        emb_dataset = f.create_dataset("image_embs", shape=(total_images, EMB_DIM), dtype=np.float32)

        model = CLIP_Wrapper(device=DEVICE)

        metadata_rows = []
        offset = 0

        for subset, split, ds in datasets_all:
            for i in tqdm(range(0, len(ds), batch_size), desc=f"Encoding {subset}-{split}"):
                batch = ds[i:i+batch_size]

                # ✅ Fix: batch is dict-of-lists, not list-of-dicts
                pil_imgs = batch["image"]

                # GPU encode
                emb = model.encode_images(pil_imgs)
                emb = normalize(emb)

                start = offset + i
                end = start + len(pil_imgs)
                emb_dataset[start:end, :] = emb

                # Save metadata for each image in the batch
                for j in range(len(pil_imgs)):
                    metadata_rows.append({
                        "subset": subset,
                        "split": split,
                        "index_in_split": i + j,
                        "file_path": (batch.get("file", [None]*len(pil_imgs))[j]
                                      or batch.get("image_path", [None]*len(pil_imgs))[j]),
                        "age": batch.get("age", [None]*len(pil_imgs))[j],
                        "gender": batch.get("gender", [None]*len(pil_imgs))[j],
                        "race": batch.get("race", [None]*len(pil_imgs))[j],
                    })

            offset += len(ds)

    # Save metadata CSV
    pd.DataFrame(metadata_rows).to_csv(metadata_path, index=False)

    print(f"✅ Saved embeddings -> {embeddings_path}")
    print(f"✅ Saved metadata  -> {metadata_path}")

def main():
    stream_and_save_embeddings()

if __name__ == "__main__":
    main()

# ==============================
# PART 2: Re-run per profession
# ==============================

import numpy as np
import pandas as pd
from pathlib import Path
from transformers import CLIPProcessor, CLIPModel
import torch

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---- Config ----
OUTPUT_DIR = Path("output_openai_clip")
embeddings_path = OUTPUT_DIR / "fairface_clip_embeddings.h5"   # HDF5 embeddings
metadata_path = OUTPUT_DIR / "fairface_metadata.csv"

DEVICE = "cuda:0"
TOP_K = 100

# ---- Change this manually each run ----
PROFESSION = "Midwife"
PROMPT = f"Photo of a {PROFESSION}"
EXCEL_OUTPUT = OUTPUT_DIR / f"{PROFESSION.replace(' ', '_')}.xlsx"

# ---- CLIP wrapper for text ----
class OpenAI_CLIP_Wrapper:
    def __init__(self, model_name="openai/clip-vit-base-patch16", device="cuda"):
        self.device = device
        self.model = CLIPModel.from_pretrained(model_name).to(device)
        self.processor = CLIPProcessor.from_pretrained(model_name)
        self.model.eval()
    def encode_texts(self, texts):
        inputs = self.processor(text=texts, return_tensors="pt", padding=True).to(self.device)
        with torch.no_grad():
            text_emb = self.model.get_text_features(**inputs)
        return text_emb.cpu().numpy()

# ---- Excel output (no images, only metadata) ----
def save_to_excel(df, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = PROFESSION

    # column headers
    headers = ["Rank","Similarity","Subset","Split","Index","FilePath","Age","Gender","Race"]
    for col_i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i, value=h)
        ws.column_dimensions[get_column_letter(col_i)].width = 20

    # rows
    row_ptr = 2
    for _, r in df.iterrows():
        ws.cell(row=row_ptr, column=1, value=int(r["rank"]))
        ws.cell(row=row_ptr, column=2, value=float(r["similarity"]))
        ws.cell(row=row_ptr, column=3, value=str(r["subset"]))
        ws.cell(row=row_ptr, column=4, value=str(r["split"]))
        ws.cell(row=row_ptr, column=5, value=int(r["index_in_split"]))
        ws.cell(row=row_ptr, column=6, value=str(r.get("file_path")))
        ws.cell(row=row_ptr, column=7, value=str(r.get("age")))
        ws.cell(row=row_ptr, column=8, value=str(r.get("gender")))
        ws.cell(row=row_ptr, column=9, value=str(r.get("race")))
        row_ptr += 1

    wb.save(excel_path)
    print(f"📊 Excel saved: {excel_path}")

# ---- Main profession search ----
def main():
    import h5py
    # Load embeddings + metadata
    with h5py.File(embeddings_path, "r") as f:
        image_embs = f["image_embs"][:]

    records = pd.read_csv(metadata_path).fillna("").to_dict("records")

    # Encode profession
    model = OpenAI_CLIP_Wrapper(device=DEVICE)
    text_emb = model.encode_texts([PROMPT])[0]
    text_emb = text_emb / (np.linalg.norm(text_emb) + 1e-12)

    sims = image_embs.dot(text_emb)

    print(f"🔍 Total images checked for '{PROFESSION}': {len(sims)}")

    # top-100
    top_idx = np.argpartition(-sims, TOP_K-1)[:TOP_K]
    top_idx = top_idx[np.argsort(-sims[top_idx])]
    print(f"✅ Top {TOP_K} images selected for '{PROFESSION}'")

    rows = []
    for rank, idx in enumerate(top_idx, start=1):
        rec = records[idx]
        rec["rank"] = rank
        rec["similarity"] = float(sims[idx])
        rows.append(rec)
    df_top = pd.DataFrame(rows)

    # ---- Group summaries ----
    print("\n--- RAW AGE DISTRIBUTION ---")
    print(df_top["age"].value_counts(dropna=False))

    # collapse into Young/Adult/Old
    age_mapping = {
        "0-2": "Young", "3-9": "Young", "10-19": "Young",
        "20-29": "Adult", "30-39": "Adult", "40-49": "Adult",
        "50-59": "Old", "60-69": "Old", "more than 70": "Old"
    }
    df_top["age_group"] = df_top["age"].map(age_mapping).fillna("Unknown")

    print("\n--- COLLAPSED AGE GROUPS ---")
    print(df_top["age_group"].value_counts(dropna=False))

    print("\n--- GENDER DISTRIBUTION ---")
    print(df_top["gender"].value_counts(dropna=False))

    print("\n--- RACE DISTRIBUTION ---")
    print(df_top["race"].value_counts(dropna=False))

    # Save Excel (metadata only, plus age group column)
    save_to_excel(df_top, EXCEL_OUTPUT)

if __name__ == "__main__":
    main()